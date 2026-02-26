import json
import openpyxl
import ast
from tkinter import filedialog, END

def save_preferences(error_label, state):
    
    year = state["year"]
    month = state["month"]
    units_list = state["units_list"]
    workers_list = state["workers_list"]
    holiday_days = state["holiday_days"]
    shifts_list = state["shifts_list"]
    selected_cannot_days = state["selected_cannot_days"]
    selected_prefer_days = state["selected_prefer_days"]
    selected_units = state["selected_units"]
    selected_manual_days = state["selected_manual_days"]

    if year is None:
        error_label.config(text="Year not set. Please set year before saving.")
        return
    if month is None:
        error_label.config(text="Month not set. Please set month before saving.")
        return
    if not units_list:
        error_label.config(text="Units not set. Please set units before saving.")
        return
    
    file_path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON files", "*.json"), ("All files", "*.*")])
    if file_path:
        # Clean up the dictionaries before saving
        # Get a set of row numbers that actually have saved workers
        saved_row_numbers = set()  # Like a basket to collect valid table numbers, it gives every row_num an id, so when looking for it below "if row_num in saved_row_numbers" it finds it instantly
        for worker in workers_list:  # Loop through all saved workers
            saved_row_numbers.add(worker["worker_row_number"])  # Add their table number
        
        # Filter the dictionaries to only keep entries for saved workers
        # This is like throwing away notes from empty tables
        cleaned_cannot = {row_num: days for row_num, days in selected_cannot_days.items() if row_num in saved_row_numbers}
        cleaned_prefer = {row_num: days for row_num, days in selected_prefer_days.items() if row_num in saved_row_numbers}
        cleaned_selected_units = {row_num: days for row_num, days in selected_units.items() if row_num in saved_row_numbers}
        cleaned_manual = {row_num: days for row_num, days in selected_manual_days.items() if row_num in saved_row_numbers}
        
        # Build the data to save (using cleaned versions)
        data = {"year": year}
        data["month"] = month
        data["holiday_days"] = holiday_days
        data["shifts_list"] = shifts_list
        data["workers_list"] = workers_list
        data["selected_cannot_days"] = cleaned_cannot  # Use cleaned version
        data["selected_prefer_days"] = cleaned_prefer  # Use cleaned version
        data["selected_units"] = cleaned_selected_units
        data["selected_manual_days"] = cleaned_manual  # Use cleaned version
        data["units_list"] = units_list  # Save units_list
        
        with open(file_path, 'w') as f:
            json.dump(data, f)
        error_label.config(text="Preferences saved.")

def load_preferences(widgets, state, callbacks):
    """
    Load preferences from a JSON file and restore the full application state.

    Args:
        widgets (dict): GUI elements to update. Keys:
            - "error_label"         : status message label
            - "current_year_label"  : label showing current year
            - "year_entry"          : entry box for year
            - "month_entry"         : entry box for month
            - "current_month_label" : label showing current month
            - "holiday_entry"       : entry box for holiday days
            - "holidays_label"      : label showing holiday list
            - "units_entry"         : entry box for units
            - "current_units_label" : label showing current units

        state (dict): Mutable objects shared with the main file.
            Lists and dicts are modified in-place, so changes here
            automatically reflect in hospital_rota_app.py. Keys:
            - "workers_list"         : list of worker dictionaries
            - "worker_rows"          : list of GUI row widget dicts
            - "selected_cannot_days" : dict {row_num: [shift strings]}
            - "selected_prefer_days" : dict {row_num: [shift strings]}
            - "selected_units"       : dict {row_num: [unit strings]}
            - "selected_manual_days" : dict {row_num: [shift strings]}
            - "holiday_days"         : list of holiday day numbers
            - "units_list"           : list of unit name strings
            - "shifts_list"          : list of shift dictionaries

        callbacks (dict): Functions from the main file. Keys:
            - "save_month"            : recalculates days_list after month loads
            - "add_worker_row"        : creates a new GUI row for a worker
            - "set_worker_row_number" : sets the global worker_row_number integer

    Returns:
        dict with keys "year", "month", "worker_row_number" if a file was
        loaded successfully, or None if the user cancelled the file dialog.
        The wrapper in hospital_rota_app.py applies these back to globals.
    """

    # Unpack widgets
    error_label        = widgets["error_label"]
    current_year_label = widgets["current_year_label"]
    year_entry         = widgets["year_entry"]
    month_entry        = widgets["month_entry"]
    current_month_label = widgets["current_month_label"]
    holiday_entry      = widgets["holiday_entry"]
    holidays_label     = widgets["holidays_label"]
    units_entry        = widgets["units_entry"]
    current_units_label = widgets["current_units_label"]

    # Unpack mutable state (lists/dicts — modified in-place, no need to return them)
    workers_list        = state["workers_list"]
    worker_rows         = state["worker_rows"]
    selected_cannot_days = state["selected_cannot_days"]
    selected_prefer_days = state["selected_prefer_days"]
    selected_units      = state["selected_units"]
    selected_manual_days = state["selected_manual_days"]
    holiday_days        = state["holiday_days"]
    units_list          = state["units_list"]
    shifts_list         = state["shifts_list"]

    # Unpack callbacks
    save_month            = callbacks["save_month"]
    add_worker_row        = callbacks["add_worker_row"]
    set_worker_row_number = callbacks["set_worker_row_number"]  # sets the global integer in main file
    set_year              = callbacks["set_year"]
    set_month             = callbacks["set_month"]

    # Scalars — local only, returned at the end so the wrapper can update globals
    year             = None
    month            = None
    worker_row_number = 1

    file_path = filedialog.askopenfilename(filetypes=[("JSON files", "*.json"), ("All files", "*.*")])
    if not file_path:
        return None  # User clicked Cancel — nothing to update

    with open(file_path, 'r') as f:
        data = json.load(f)

    if "year" in data:
        year = data["year"]
        set_year(year)                          # ← add this line
        current_year_label.config(text="Current Year: " + str(year))
        year_entry.delete(0, END)
        year_entry.insert(0, str(year))
    if "month" in data:
        month = data["month"]
        set_month(month)                        # ← add this line
        month_entry.delete(0, END)
        month_entry.insert(0, str(month))
        month_name_list = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
        month_name = month_name_list[month-1]
        current_month_label.config(text="Current Month: " + month_name)
        save_month()
    if "holiday_days" in data:
        holiday_days.clear()
        holiday_days.extend(data["holiday_days"])
        holiday_entry.delete(0, END)
        if holiday_days:
            holiday_entry.insert(0, ", ".join(map(str, holiday_days)))
            holidays_label.config(text="Holiday List: " + str(holiday_days))
        else:
            holidays_label.config(text="Holiday List: None")
    if "units_list" in data:
        units_list.clear()
        units_list.extend(data["units_list"])
        nice_units_text = ", ".join(units_list)
        units_entry.delete(0, END)
        if units_list:
            units_entry.insert(0, nice_units_text)
            current_units_label.config(text=f"Current Units: {', '.join(units_list)}")
        else:
            current_units_label.config(text="Current Units: None")
    if "shifts_list" in data:
        shifts_list.clear()
        shifts_list.extend(data["shifts_list"])
    if "workers_list" in data:
        workers_list.clear()
        workers_list.extend(data["workers_list"])
        if "selected_cannot_days" not in data:
            selected_cannot_days.clear()
            selected_cannot_days.update({worker["worker_row_number"]: worker.get("cannot_work", []) for worker in workers_list})
        # Clear existing worker rows
        for row_widgets in worker_rows:
            for key, widget in row_widgets.items():
                if key != 'row_num' and widget.winfo_exists():
                    widget.destroy()
        worker_rows.clear()
        worker_row_number = 1
        # Now add rows for loaded workers
        for worker in workers_list:
            row_num = worker["worker_row_number"]
            set_worker_row_number(row_num)  # sets global in main file so add_worker_row uses correct number
            add_worker_row()
        # Now populate the entries
        for worker in workers_list:
            row_num = worker["worker_row_number"]
            for rw in worker_rows:
                if rw['row_num'] == row_num:
                    rw['name_entry'].delete(0, END)
                    rw['name_entry'].insert(0, worker["name"])
                    rw['range_entry'].delete(0, END)
                    rw['range_entry'].insert(0, f"{worker['shifts_to_fill'][0]}-{worker['shifts_to_fill'][1]}")
                    rw['max_weekends_entry'].delete(0, END)
                    rw['max_weekends_entry'].insert(0, str(worker["max_weekends"]))
                    rw['max_24hr_entry'].delete(0, END)
                    rw['max_24hr_entry'].insert(0, str(worker["max_24hr"]))
                    break
        # Update worker_row_number to max +1
        if workers_list:
            max_row = max(worker["worker_row_number"] for worker in workers_list)
            worker_row_number = max_row + 1
        else:
            worker_row_number = 1
    if "selected_cannot_days" in data:
        selected_cannot_days.clear()
        selected_cannot_days.update({int(k): v for k, v in data["selected_cannot_days"].items()})
    if "selected_prefer_days" in data:
        selected_prefer_days.clear()
        selected_prefer_days.update({int(k): v for k, v in data["selected_prefer_days"].items()})
    if "selected_units" in data:
        selected_units.clear()
        selected_units.update({int(k): v for k, v in data["selected_units"].items()})
    if "selected_manual_days" in data:
        selected_manual_days.clear()
        selected_manual_days.update({int(k): v for k, v in data["selected_manual_days"].items()})
    # Set worker dicts from loaded selected dicts to ensure consistency
    for worker in workers_list:
        row_num = worker["worker_row_number"]
        worker["cannot_work"] = selected_cannot_days.get(row_num, [])
        worker["prefers"] = selected_prefer_days.get(row_num, [])
        worker["prefer_units"] = selected_units.get(row_num, [])
        if row_num not in selected_manual_days:
            selected_manual_days[row_num] = []
    error_label.config(text="Preferences loaded.")
    # Update button texts to show selections
    for rw in worker_rows:
        row_num = rw['row_num']
        num_cannot = len(selected_cannot_days.get(row_num, []))
        rw['cannot_button'].config(text=f"Select ({num_cannot})" if num_cannot > 0 else "Select")
        num_prefer = len(selected_prefer_days.get(row_num, []))
        rw['prefer_button'].config(text=f"Select ({num_prefer})" if num_prefer > 0 else "Select")
        num_prefer_unit = len(selected_units.get(row_num, []))
        rw['prefer_unit_button'].config(text=f"Select ({num_prefer_unit})" if num_prefer_unit > 0 else "Select")
        num_manual = len(selected_manual_days.get(row_num, []))
        rw['manual_button'].config(text=f"Select ({num_manual})" if num_manual > 0 else "Select")

    print(f"Debugging. Year: {year}, month {month}, Holiday list: {holiday_days}, Shift list length {len(shifts_list)}")
    for worker in workers_list:
        print(worker)
    print(selected_cannot_days)

    # Return scalars so the wrapper in hospital_rota_app.py can update its globals
    return {"year": year, "month": month, "worker_row_number": worker_row_number}

def load_xlsx_preferences(widgets, state, callbacks):
    """
    Load worker preferences from an Excel (.xlsx) file.

    Expected format:
    - Row 1: Headers with "Name" in column A, day numbers (1, 2, 3...) in columns B onwards
    - Row 2+: Worker names in column A, colored cells for preferences
    - RED cells = Cannot work that day (both Day & Night shifts)
    - GREEN cells = Prefer to work that day (both Day & Night shifts)

    Args:
        widgets (dict): GUI elements to update. Keys:
            - "error_label"         : status message label
            - "units_entry"         : entry box for units
            - "current_units_label" : label showing current units
            - "root"                : main window, needed to force GUI refresh

        state (dict): Mutable objects modified in-place. Keys:
            - "workers_list"         : list of worker dictionaries
            - "worker_rows"          : list of GUI row widget dicts
            - "selected_cannot_days" : dict {row_num: [shift strings]}
            - "selected_prefer_days" : dict {row_num: [shift strings]}
            - "selected_manual_days" : dict {row_num: [shift strings]}
            - "selected_units"       : dict {row_num: [unit strings]}
            - "units_list"           : list of unit name strings
            - "days_list"            : list of days in the month (read-only)

        callbacks (dict): Functions from the main file. Keys:
            - "add_worker_row"        : creates a new GUI row for a worker
            - "make_shifts"           : rebuilds the shifts list after loading
            - "set_worker_row_number" : sets the global worker_row_number integer

    Returns:
        dict with key "worker_row_number" on success, or None if cancelled/error.
        The wrapper in hospital_rota_app.py applies this back to the global.
    """
    # Unpack widgets
    error_label         = widgets["error_label"]
    units_entry         = widgets["units_entry"]
    current_units_label = widgets["current_units_label"]
    root                = widgets["root"]

    # Unpack mutable state (modified in-place)
    workers_list         = state["workers_list"]
    worker_rows          = state["worker_rows"]
    selected_cannot_days = state["selected_cannot_days"]
    selected_prefer_days = state["selected_prefer_days"]
    selected_manual_days = state["selected_manual_days"]
    selected_units       = state["selected_units"]
    units_list           = state["units_list"]
    days_list            = state["days_list"]

    # Unpack callbacks
    add_worker_row        = callbacks["add_worker_row"]
    make_shifts           = callbacks["make_shifts"]
    set_worker_row_number = callbacks["set_worker_row_number"]

    # IMPORTANT: First check if days_list exists (year and month must be set)
    try:
        if not days_list:
            error_label.config(text="Error: Please set year and month first!")
            return None
    except NameError:
        error_label.config(text="Error: Please set year and month first!")
        return None

    # Open file dialog
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
    )
    if not file_path:
        return None  # User clicked Cancel

    # Local counter to track worker_row_number — kept in sync with the global
    # via set_worker_row_number() and add_worker_row()
    local_wrn = 1

    try:
        # data_only=False → keep formulas and formatting (needed to read cell colors)
        # data_only=True would only give us the VALUES, losing all color information
        wb = openpyxl.load_workbook(file_path, data_only=False)

        # Try to find a suitable sheet
        sheet = None

        # If only one sheet, just use it
        if len(wb.sheetnames) == 1:
            sheet = wb[wb.sheetnames[0]]
        else:
            # Try to find sheet by name (supports multiple languages)
            for sname in wb.sheetnames:
                sname_lower = sname.lower()
                if any(keyword in sname_lower for keyword in
                       ["feuille", "sheet", "preferences", "rota", "workers", "staff"]):
                    sheet = wb[sname]
                    break

        # Fallback: use active sheet
        if sheet is None:
            sheet = wb.active

        # STEP: Read units from Row 1
        # Scan every cell in Row 1 looking for one that starts with "Units_list=["
        units_col_val  = None
        prefer_unit_col = None

        for col in range(1, sheet.max_column + 1):
            cell_val = sheet.cell(row=1, column=col).value
            if cell_val and str(cell_val).strip().startswith("Units_list=["):
                units_col_val   = str(cell_val).strip()
                prefer_unit_col = col
                break

        if units_col_val is None:
            error_label.config(text="Error: Could not find 'Units_list=[...]' in Row 1. Please add it to the xlsx.")
            return None

        # Parse the list out of the string e.g. 'Units_list=["Cardiology", "Internal Medicine"]'
        bracket_start = units_col_val.index('[')
        bracket_part  = units_col_val[bracket_start:]
        loaded_units  = ast.literal_eval(bracket_part)

        # Update units_list in-place (clear + extend instead of reassigning)
        units_list.clear()
        units_list.extend(loaded_units)
        units_entry.delete(0, END)
        units_entry.insert(0, ", ".join(units_list))
        current_units_label.config(text=f"Current Units: {', '.join(units_list)}")
        error_label.config(text=f"Units loaded: {', '.join(units_list)}")

        # Find header row and the "Name" column
        # HEADER_ROW stays 1 — that's where day numbers and units always live
        HEADER_ROW = 1

        # But "Name" might not be in Row 1 — scan column 1 to find it
        name_col = 1
        NAME_ROW = None

        for row in range(1, sheet.max_row + 1):
            val = sheet.cell(row=row, column=1).value
            if val and str(val).strip().lower() == "name":
                NAME_ROW = row
                break

        if NAME_ROW is None:
            error_label.config(text="Error: Could not find 'Name' in column 1.")
            return None

        # Worker data always starts on the row immediately below "Name"
        FIRST_DATA_ROW = NAME_ROW + 1

        # Find the Shift_range column by scanning Row 1
        shift_range_col = None
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=1, column=col).value
            if val and str(val).strip().lower() == "shift_range":
                shift_range_col = col
                break

        # Maps each column to its specific (day_number, shift_type) pair
        # by reading Row 1 (day numbers) AND Row 2 (D or N labels) together
        col_to_shift = {}
        current_day  = None
        SHIFT_TYPE_ROW = 2

        for col in range(1, sheet.max_column + 1):
            day_val = sheet.cell(row=HEADER_ROW,    column=col).value
            dn_val  = sheet.cell(row=SHIFT_TYPE_ROW, column=col).value

            if day_val is not None:
                try:
                    current_day = int(float(day_val))
                except (ValueError, TypeError):
                    current_day = None

            if current_day is None:
                continue
            if not (1 <= current_day <= len(days_list)):
                continue
            if dn_val not in ("D", "N"):
                continue

            shift_type = "Day" if dn_val == "D" else "Night"
            col_to_shift[col] = (current_day, shift_type)
            
            """
            col_to_shift is a dictionary that maps a column number → (day, shift_type) tuple.
            For example, if column 3 in the xlsx corresponds to Day 5, Night shift, then after this line:
            col_to_shift[3] = (5, "Night")
            """

        if not col_to_shift:
            error_label.config(text="Error: Could not find D/N shift columns in Row 2.")
            return None

        # Clear old data completely
        for row_widgets in worker_rows[:]:  # [:] copies the list so we can safely clear it
            for key, widget in row_widgets.items():
                if key != 'row_num' and widget.winfo_exists():
                    widget.destroy()

        worker_rows.clear()
        workers_list.clear()
        selected_cannot_days.clear()
        selected_prefer_days.clear()
        selected_manual_days.clear()

        # Reset worker_row_number to 1 in main file; track locally with local_wrn
        set_worker_row_number(1)
        local_wrn = 1

        loaded_count = 0
        total_rows = sheet.max_row - FIRST_DATA_ROW + 1

        # Loop through each worker row
        for row_idx, r in enumerate(range(FIRST_DATA_ROW, sheet.max_row + 1), 1):
            name_val = sheet.cell(row=r, column=name_col).value

            if not name_val:
                continue

            name = str(name_val).strip()
                                                                # skips if the name is an empty string "" (after .strip()
            if not name or not any(c.isalpha() for c in name): # skips if the name has no alphabetic characters at all, e.g. a row that's just "123" or "---" or whitespace that slipped through
                continue

            # Collect cannot/prefer days by checking cell colors
            cannot_list = []
            prefer_list = []

            for col, (day, shift_type) in col_to_shift.items():
                cell = sheet.cell(row=r, column=col)

                if cell.fill and cell.fill.fill_type == 'solid':
                    color = cell.fill.start_color.rgb
                    if color is None:
                        continue
                    color_hex = color.upper()[-6:]

                    if is_red_color(color_hex):
                        cannot_list.append(f"{shift_type} {day}")
                    elif is_green_color(color_hex):
                        prefer_list.append(f"{shift_type} {day}")

            # Read preferred units for this worker
            prefer_units = []

            if prefer_unit_col is not None:
                raw = sheet.cell(row=r, column=prefer_unit_col).value

                if raw is not None:
                    if isinstance(raw, float):
                        parts = str(raw).strip().split(".")   # "1.2" → ["1", "2"]
                    else:
                        parts = str(raw).strip().split(",")   # "1,2" → ["1", "2"]

                    for part in parts:
                        try:
                            index = int(float(part.strip())) - 1
                            if 0 <= index < len(units_list):
                                prefer_units.append(units_list[index])
                        except (ValueError, TypeError):
                            pass

            # Read shift range for this worker
            min_shifts = 1
            max_shifts = 4

            if shift_range_col is not None:
                raw = sheet.cell(row=r, column=shift_range_col).value

                if raw is not None:
                    if isinstance(raw, float):
                        parts = str(raw).strip().split(".")
                    else:
                        parts = str(raw).strip().split(",")

                    if len(parts) == 2:
                        try:
                            min_shifts = int(parts[0])
                            max_shifts = int(parts[1])
                        except ValueError:
                            pass

            # Create the worker dictionary
            worker_dict = {
                "name": name,
                "shifts_to_fill": [min_shifts, max_shifts],
                "cannot_work": cannot_list,
                "prefers": prefer_list,
                "prefer_units": prefer_units,
                "max_weekends": 100,
                "max_24hr": 100,
                "worker_row_number": local_wrn
            }
            workers_list.append(worker_dict)

            current_row_num = local_wrn

            add_worker_row()  # increments global worker_row_number by 1
            local_wrn += 1    # keep local counter in sync

            # Fill in the values in the GUI
            for row_widgets in worker_rows:
                if row_widgets['row_num'] == current_row_num:
                    row_widgets['name_entry'].delete(0, END)
                    row_widgets['name_entry'].insert(0, name)

                    row_widgets['range_entry'].delete(0, END)
                    row_widgets['range_entry'].insert(0, f"{min_shifts}-{max_shifts}")

                    row_widgets['max_weekends_entry'].delete(0, END)
                    row_widgets['max_weekends_entry'].insert(0, "100")

                    row_widgets['max_24hr_entry'].delete(0, END)
                    row_widgets['max_24hr_entry'].insert(0, "100")

                    row_widgets['cannot_button'].config(
                        text=f"Select ({len(cannot_list)})" if cannot_list else "Select"
                    )
                    row_widgets['prefer_button'].config(
                        text=f"Select ({len(prefer_list)})" if prefer_list else "Select"
                    )
                    row_widgets['prefer_unit_button'].config(
                        text=f"Select ({len(prefer_units)})" if prefer_units else "Select"
                    )
                    row_widgets['manual_button'].config(text="Select")
                    break

            # Store selections for the popup windows
            selected_cannot_days[current_row_num] = cannot_list
            selected_prefer_days[current_row_num] = prefer_list
            selected_units[current_row_num]       = prefer_units
            selected_manual_days[current_row_num] = []

            loaded_count += 1

            # Show progress every 5 workers (helps with large files)
            if loaded_count % 5 == 0 or row_idx == total_rows:
                error_label.config(text=f"Loading... {loaded_count} worker(s)")
                root.update()  # Force GUI to refresh

        # Final success message
        error_label.config(text=f"✓ Loaded {loaded_count} worker(s) — red = cannot, green = prefer")

    except Exception as e:
        error_label.config(text=f"Error reading file: {str(e)}")
        print("Detailed error:", e)
        return None

    return {"worker_row_number": local_wrn}

# Helper functions for color detection
def is_red_color(color_hex):
    """
    Check if a color hex code (RRGGBB) represents red.
    Returns True for pure red and close variations.
    
    Args:
        color_hex: 6-character hex string like "FF0000"
    
    Returns:
        bool: True if the color is red-ish
    """
    try:
        r = int(color_hex[0:2], 16)
        g = int(color_hex[2:4], 16)
        b = int(color_hex[4:6], 16)
        # Red if: red channel is the biggest AND clearly dominates green and blue
        return r > g and r > b and r > 100
    except (ValueError, IndexError):
        return False


def is_green_color(color_hex):
    """
    Check if a color hex code (RRGGBB) represents green.
    Returns True for pure green and close variations.
    
    Args:
        color_hex: 6-character hex string like "00FF00"
    
    Returns:
        bool: True if the color is green-ish
    """
    try:
        r = int(color_hex[0:2], 16)
        g = int(color_hex[2:4], 16)
        b = int(color_hex[4:6], 16)
        # Green if: green channel is the biggest AND clearly dominates red and blue
        return g > r and g > b and g > 100
    except (ValueError, IndexError):
        return False
        